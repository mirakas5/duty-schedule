"""엑셀(.xlsx)에서 인원 이름을 추출 — Design Ref: §2.3 Infra(excel).

이름 컬럼 자동 탐지: 헤더에 '이름/성명/name' 포함 컬럼 우선, 없으면 첫 컬럼.
"""
from __future__ import annotations

import io
from typing import List

from openpyxl import load_workbook

_NAME_HINTS = ("이름", "성명", "name", "직원", "사원")


def extract_names(content: bytes) -> List[str]:
    """xlsx 바이트에서 이름 목록을 추출(공백 제거·빈값 제외·중복 제거, 순서 유지)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # 헤더에서 이름 컬럼 인덱스 탐지
    header = rows[0]
    name_col = 0
    has_header = False
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        text = str(cell).strip().lower()
        if any(h in text for h in (_h.lower() for _h in _NAME_HINTS)):
            name_col = idx
            has_header = True
            break

    data_rows = rows[1:] if has_header else rows

    seen = set()
    names: List[str] = []
    for row in data_rows:
        if name_col >= len(row):
            continue
        val = row[name_col]
        if val is None:
            continue
        name = str(val).strip()
        if not name or name in seen:
            continue
        seen.add(name)
        names.append(name)
    return names
